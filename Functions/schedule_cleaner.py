import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from dotenv import load_dotenv
from logger import coach_logger
import os

load_dotenv()

class CFClassesDataCleaner:
    def __init__(self, orginal_input_file_path, date_preserved_file_path):
        self.output_path = orginal_input_file_path
        self.date_preserved_file_path = date_preserved_file_path
        self.date_row = None
        self.time_column = None
        self.cf_classes_df_partial = None

    def preserve_dates(self, sheet_name='CF Classes', date_row_marker='Date ->'):
        coach_logger.log_info(f'[+] Pre Cleaning {os.path.basename(self.output_path)}...')
        original_wb_with_formula = load_workbook(self.output_path, data_only=False)
        original_ws_with_formula = original_wb_with_formula[sheet_name]
        
        original_wb_with_values = load_workbook(self.output_path, data_only=True)
        original_ws_with_values = original_wb_with_values[sheet_name]
        
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        date_row_index = None

        for idx, row in enumerate(original_ws_with_values.iter_rows(values_only=True)):
            if row[2] == date_row_marker:
                date_row_index = idx
            if date_row_index is not None and idx == date_row_index + 1:
                date_row_with_formula = next(original_ws_with_formula.iter_rows(min_row=idx + 1, max_row=idx + 1, values_only=False))
                date_row_with_values = next(original_ws_with_values.iter_rows(min_row=idx + 1, max_row=idx + 1, values_only=True))
                new_ws.append([cell_value if cell_value is not None else cell_formula.value for cell_value, cell_formula in zip(date_row_with_values, date_row_with_formula)])
            else:
                new_ws.append(row)

        new_wb.save(self.date_preserved_file_path)
        coach_logger.log_info(f'[+] Pre Cleaned {os.path.basename(self.output_path)}')
        coach_logger.log_info(f'[+] Date preserved file saved to {os.path.basename(self.date_preserved_file_path)}')

    def load_data(self):
        coach_logger.log_info(f'[+] Loading data from {os.path.basename(self.date_preserved_file_path)}')
        cf_classes_df_debug = pd.read_excel(self.date_preserved_file_path, sheet_name='CF Classes')
        date_row_index = cf_classes_df_debug[cf_classes_df_debug['Unnamed: 2'] == 'Date ->'].index[0]
        self.date_row = cf_classes_df_debug.iloc[date_row_index, 3:].dropna()
        self.cf_classes_df_partial = cf_classes_df_debug.iloc[6:21, 2:]
        self.time_column = self.cf_classes_df_partial['Unnamed: 2'].dropna()

    
    def set_date_as_header(self):
        coach_logger.log_info(f'[+] Formatting final output...')
        # Store a copy of the original 'Unnamed: 2' column for preservation
        original_unnamed_2 = self.cf_classes_df_partial['Unnamed: 2'].copy()
        
        formatted_dates = self.format_datetime_to_mdy(self.date_row.tolist())
        
        week_separator_counter = 0
        insert_week_separator = False
        new_columns = ['Unnamed: 2']

        for i, col in enumerate(self.cf_classes_df_partial.columns[1:]):
            if insert_week_separator and formatted_dates:
                next_date_str = formatted_dates[0]
                if next_date_str != '01/02/2023':
                    new_columns.append(f"Week_Separator_{week_separator_counter}")
                    week_separator_counter += 1
                insert_week_separator = False

            if formatted_dates:
                date_str = formatted_dates.pop(0)
                new_columns.append(date_str)
                if pd.Timestamp(date_str).day_name() == 'Sunday':
                    insert_week_separator = True

        # Update the columns of the original DataFrame
        self.cf_classes_df_partial.columns = new_columns

        # Restore the original 'Unnamed: 2' column
        self.cf_classes_df_partial['Unnamed: 2'] = original_unnamed_2

    def get_date_row(self):
        return self.date_row

    def get_time_column(self):
        return self.time_column

    def get_cf_classes_df_partial(self):
        return self.cf_classes_df_partial
    
    @staticmethod
    def format_datetime_to_mdy(datetime_objs):
        return [datetime_obj.strftime('%m/%d/%Y') for datetime_obj in datetime_objs]
    


def main():
    orginal_input_file_path = os.getenv('INPUT_FILE_PATH')
    date_preserved_file_path = os.getenv('PRESERVED_OUTPUT_FILE_PATH')
    clened_file_path = os.getenv('CLEANED_OUTPUT_FILE_PATH')
    cf_classes_cleaner = CFClassesDataCleaner(orginal_input_file_path, date_preserved_file_path)
    cf_classes_cleaner.preserve_dates()
    cf_classes_cleaner.load_data()
    cf_classes_cleaner.set_date_as_header()
    result_df = cf_classes_cleaner.get_cf_classes_df_partial()
    result_df.to_excel(clened_file_path, index=False)

if __name__ == '__main__':
    main()

# cf_classes_cleaner = CFClassesDataCleaner(orginal_input_file_path, date_preserved_file_path)
# cf_classes_cleaner.preserve_dates()
# cf_classes_cleaner.load_data()
# cf_classes_cleaner.set_date_as_header()
# result_df = cf_classes_cleaner.get_cf_classes_df_partial()
# result_df.to_excel(clened_file_path, index=False)